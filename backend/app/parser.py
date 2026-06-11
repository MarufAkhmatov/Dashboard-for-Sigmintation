"""Multi-format Jira export parser: CSV / XLSX / HTML -> list[dict] of raw rows."""
import csv
import io
from pathlib import Path


def parse_file(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    data = path.read_bytes()
    if suffix == ".csv":
        return _parse_csv(data)
    if suffix in (".xlsx", ".xlsm"):
        return _parse_xlsx(data)
    if suffix in (".html", ".htm"):
        return _parse_html(data)
    # try CSV as a fallback
    return _parse_csv(data)


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _parse_csv(data: bytes) -> list[dict]:
    text = _decode(data)
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except Exception:
        pass
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return []
    header = [h.strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c.strip() for c in r):
            continue
        # Jira CSV can repeat columns (e.g. multiple "Status"/"Outward issue link");
        # collapse duplicates into a comma-joined string.
        rec: dict[str, str] = {}
        for i, col in enumerate(header):
            val = r[i].strip() if i < len(r) else ""
            if col in rec and val:
                rec[col] = f"{rec[col]}|{val}" if rec[col] else val
            else:
                rec.setdefault(col, val)
        out.append(rec)
    return out


def _parse_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or not any(c is not None and str(c).strip() for c in r):
            continue
        rec = {}
        for i, col in enumerate(header):
            v = r[i] if i < len(r) else None
            rec[col] = "" if v is None else str(v).strip()
        out.append(rec)
    return out


def _parse_html(data: bytes) -> list[dict]:
    text = _decode(data)
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(text, "html.parser")
    except Exception:
        return _parse_html_stdlib(text)
    table = soup.find("table")
    if not table:
        return []
    rows = table.find_all("tr")
    header_cells = rows[0].find_all(["th", "td"])
    header = [c.get_text(strip=True) for c in header_cells]
    out = []
    for tr in rows[1:]:
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        rec = {}
        for i, col in enumerate(header):
            rec[col] = cells[i].get_text(strip=True) if i < len(cells) else ""
        if any(rec.values()):
            out.append(rec)
    return out


def _parse_html_stdlib(text: str) -> list[dict]:
    from html.parser import HTMLParser

    class T(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows, self.cur, self.cell, self.in_cell = [], None, "", False

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self.cur = []
            elif tag in ("td", "th"):
                self.in_cell, self.cell = True, ""

        def handle_endtag(self, tag):
            if tag in ("td", "th") and self.cur is not None:
                self.cur.append(self.cell.strip())
                self.in_cell = False
            elif tag == "tr" and self.cur is not None:
                self.rows.append(self.cur)
                self.cur = None

        def handle_data(self, data):
            if self.in_cell:
                self.cell += data

    p = T()
    p.feed(text)
    if not p.rows:
        return []
    header = p.rows[0]
    out = []
    for r in p.rows[1:]:
        rec = {header[i]: (r[i] if i < len(r) else "") for i in range(len(header))}
        if any(rec.values()):
            out.append(rec)
    return out
